#!/usr/bin/env python3
"""
MONDIALITO 2026 — Generatore Excel Pronostici
==============================================
Uso:
    python genera-excel.py [pronostici_mondialito_YYYY-MM-DD.json]

Se non passi il file JSON, cerca automaticamente il file più recente
nella stessa cartella.

Output: Mondialito2026_Pronostici.xlsx

Fogli generati:
  RISULTATI      — input giallo per i risultati ufficiali (gironi + elim)
  GIRONI         — pronostici gironi con formule punti live
  GRIGLIA        — pronostici posizioni finali con formule punti live
  ELIMINATORIE   — pronostici fase KO (valori pre-calcolati)
  CAPOCANNONIERE — pronostici top scorer (valori pre-calcolati)
  CLASSIFICA     — riepilogo finale con SUM dalle sheets live

Dipendenza:  pip install openpyxl
"""

import json
import sys
import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
except ImportError:
    print("❌ openpyxl non installato. Esegui: pip install openpyxl")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
# STILI
# ══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=11, italic=False, name='Calibri'):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style='thin', color='BBBBBB'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_full(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

# Palette colori
BLU_SCURO    = '1A3A5C'
BLU_MEDIO    = '2E6DA4'
BLU_CHIARO   = 'CDDFF5'
VERDE_SCURO  = '375623'
VERDE        = '70AD47'
VERDE_CHIARO = 'E2EFDA'
ARANCIO      = 'C55A11'
ARANCIO_CH   = 'FCE4D6'
GIALLO       = 'FFF2CC'
GIALLO_SCURO = 'F4B942'
GRIGIO       = 'D9D9D9'
GRIGIO_CH    = 'F7F7F7'
BIANCO       = 'FFFFFF'
ORO          = 'FFD700'
ROSSO        = 'C00000'
VIOLA        = '7030A0'

# Stili predefiniti
H_TITLE  = {'fill': _fill(BLU_SCURO),  'font': _font(True, BIANCO, 14), 'align': _align('center')}
H_PHASE  = {'fill': _fill(BLU_MEDIO),  'font': _font(True, BIANCO, 11), 'align': _align('center')}
H_COL    = {'fill': _fill(BLU_CHIARO), 'font': _font(True, BLU_SCURO, 10), 'align': _align('center')}
H_USER   = {'fill': _fill(ARANCIO),    'font': _font(True, BIANCO, 10),    'align': _align('center')}
H_USER2  = {'fill': _fill(ARANCIO_CH), 'font': _font(True, ARANCIO, 9),    'align': _align('center')}

INPUT_CELL = {'fill': _fill(GIALLO),      'font': _font(size=11), 'align': _align('center')}
FORMULA_CELL = {'fill': _fill(VERDE_CHIARO), 'font': _font(size=10), 'align': _align('center')}
NORM_CELL  = {'fill': _fill(BIANCO),      'font': _font(size=10), 'align': _align('center')}
LABEL_CELL = {'fill': _fill(GRIGIO_CH),   'font': _font(True, '444444', 10), 'align': _align('left')}
ZERO_CELL  = {'fill': _fill(GRIGIO_CH),   'font': _font(color='888888', size=10), 'align': _align('center')}
PT_CELL    = {'fill': _fill(VERDE_CHIARO),'font': _font(True, VERDE_SCURO, 10), 'align': _align('center')}
PT_ZERO    = {'fill': _fill(GRIGIO_CH),   'font': _font(color='AAAAAA', size=9), 'align': _align('center')}

def _style(ws, cell, style_dict):
    for attr, val in style_dict.items():
        if attr == 'fill':  cell.fill = val
        elif attr == 'font':  cell.font = val
        elif attr == 'align': cell.alignment = val
        elif attr == 'border':cell.border = val
    return cell

def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def _set_row_height(ws, row_idx, height):
    ws.row_dimensions[row_idx].height = height

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(
        start_row=r1, start_column=c1,
        end_row=r2,   end_column=c2
    )

# ══════════════════════════════════════════════════════════════════════════════
# LOGICA DI CALCOLO (stessa di functions/punteggi.js)
# ══════════════════════════════════════════════════════════════════════════════

def _segno(gc, gt):
    if gc is None or gt is None: return None
    try:
        gc, gt = int(gc), int(gt)
        return '1' if gc > gt else ('2' if gc < gt else 'X')
    except (TypeError, ValueError):
        return None

def _calcola_gironi(prono, risultati, db):
    """Ritorna { match_id: { segno_ok, esatto_ok, punti } }"""
    rg  = (risultati.get('gironi') or {})
    pg  = (prono.get('gironi') or {})
    out = {}
    for girone in db['gironi'].values():
        for p in girone['partite']:
            mid = p['id']
            r   = rg.get(mid, {})
            pr  = pg.get(mid, {})
            gc_r = r.get('gol_casa');  gt_r = r.get('gol_trasferta')
            gc_p = pr.get('gol_casa'); gt_p = pr.get('gol_trasferta')
            segno_r = _segno(gc_r, gt_r)
            segno_p = pr.get('segno')
            segno_ok = (segno_r is not None) and (segno_p == segno_r)
            esatto_ok = segno_ok and (gc_p == gc_r) and (gt_p == gt_r)
            punti = (10 if segno_ok else 0) + (5 if esatto_ok else 0)
            out[mid] = {
                'gc_r': gc_r, 'gt_r': gt_r, 'segno_r': segno_r,
                'gc_p': gc_p, 'gt_p': gt_p, 'segno_p': segno_p,
                'segno_ok': segno_ok, 'esatto_ok': esatto_ok,
                'punti': punti,
            }
    return out

def _derive_posizioni_girone(prono, db):
    """Calcola le posizioni previste nel girone partendo dai pronostici delle partite.
    Usato come fallback se posizioni_girone non è salvato nei pronostici."""
    gironi_prono = prono.get('gironi') or {}
    result = {}
    for lettera, girone in db['gironi'].items():
        standings = {sq: {'pts': 0, 'gd': 0, 'gf': 0} for sq in girone['squadre']}
        for partita in girone['partite']:
            p = gironi_prono.get(partita['id'], {}) or {}
            try:
                gc = int(p['gol_casa'])
                gt = int(p['gol_trasferta'])
            except (KeyError, TypeError, ValueError):
                continue
            casa = partita['casa']; trasf = partita['trasferta']
            if gc > gt:   standings[casa]['pts'] += 3
            elif gc < gt: standings[trasf]['pts'] += 3
            else:         standings[casa]['pts'] += 1; standings[trasf]['pts'] += 1
            standings[casa]['gd']  += gc - gt; standings[trasf]['gd']  += gt - gc
            standings[casa]['gf']  += gc;      standings[trasf]['gf']  += gt
        result[lettera] = sorted(
            girone['squadre'],
            key=lambda sq: (-standings[sq]['pts'], -standings[sq]['gd'], -standings[sq]['gf'], sq)
        )
    return result


def _calcola_griglia(prono, risultati, db):
    """Ritorna { girone: { team: { pos_prono, pos_reale, ok, punti } } }"""
    rg   = (risultati.get('posizioni_finali_gironi') or {})
    pg   = (prono.get('posizioni_girone') or {}) or _derive_posizioni_girone(prono, db)
    rElim = (risultati.get('fase_eliminatoria') or {})
    # squadre ai sedicesimi (se i risultati ci sono)
    sed_teams = set()
    sed_data  = rElim.get('sedicesimi') or {}
    for m in sed_data.values():
        if isinstance(m, dict):
            for f in ('casa', 'trasferta', 'vincitore'):
                if m.get(f): sed_teams.add(m[f])

    out = {}
    for lettera, girone in db['gironi'].items():
        pos_reale = rg.get(lettera, [])      # ['MEX','RSA','KOR','CZE']
        pos_prono = pg.get(lettera, [])      # idem, da pronostici
        d = {}
        for sq in girone['squadre']:
            pr = pos_prono.index(sq)+1 if sq in pos_prono else None
            re = pos_reale.index(sq)+1  if sq in pos_reale  else None
            # Punti solo se la squadra è avanzata ai sedicesimi (o sed_teams vuoto = non ancora)
            ok = (pr is not None and re is not None and pr == re)
            can_score = not sed_teams or sq in sed_teams
            punti = 10 if (ok and can_score) else 0
            d[sq] = {'pos_prono': pr, 'pos_reale': re, 'ok': ok, 'punti': punti}
        out[lettera] = d
    return out

def _calcola_eliminatorie(prono, risultati, db):
    """Ritorna { fase: { match_id: { vin_prono, vin_reale, mod_prono, mod_reale, sq_ok, mod_ok, punti } } }"""
    fasi = [
        ('sedicesimi', 5),
        ('ottavi', 10),
        ('quarti', 20),
        ('semifinali', 30),
    ]
    rElim = (risultati.get('fase_eliminatoria') or {})
    pElim = (prono.get('fase_eliminatoria') or {})
    out   = {}

    for fase, pti_sq in fasi:
        rFase = rElim.get(fase) or {}
        pFase = pElim.get(fase) or {}
        # set di squadre realmente in questa fase
        sq_reali = set()
        for m in rFase.values():
            if isinstance(m, dict):
                for f in ('casa', 'trasferta', 'vincitore'):
                    if m.get(f): sq_reali.add(m[f])

        fase_out = {}
        # Combina le chiavi (sia pronostici che risultati)
        tutti_match = set(list(rFase.keys()) + list(pFase.keys()))
        for mid in tutti_match:
            rM = rFase.get(mid, {}) or {}
            pM = pFase.get(mid, {}) or {}
            vp = pM.get('vincitore')
            vr = rM.get('vincitore')
            mp = pM.get('modalita')
            mr = rM.get('modalita')
            # Punti squadra: predicted vincitore era nel bracket reale?
            sq_ok  = bool(vp and sq_reali and vp in sq_reali)
            mod_ok = bool(sq_ok and vp == vr and mp and mr and mp == mr)
            punti  = (pti_sq if sq_ok else 0) + (5 if mod_ok else 0)
            fase_out[mid] = {
                'casa_r': rM.get('casa'), 'trasf_r': rM.get('trasferta'),
                'vin_prono': vp, 'vin_reale': vr,
                'mod_prono': mp, 'mod_reale': mr,
                'sq_ok': sq_ok, 'mod_ok': mod_ok, 'punti': punti,
            }
        out[fase] = fase_out

    # Finale (speciale: 2 finalisti + vincitore + modalità)
    rFinale = rElim.get('finale') or {}
    pFinale = pElim.get('finale') or {}
    # Risultato finale dal db: partita F
    rf = rFinale.get('F') or rFinale.get('partita') or {}
    pf = pFinale.get('F') or pFinale or {}
    casa_r = rf.get('casa'); trasf_r = rf.get('trasferta')
    vin_r  = rf.get('vincitore')
    mod_r  = rf.get('modalita')
    vin_p  = pf.get('vincitore') if isinstance(pf, dict) else None
    mod_p  = pf.get('modalita') if isinstance(pf, dict) else None
    # finalisti pronosticati (vincitore del match = solo finalista pronosticato)
    fin_p  = pf.get('squadre') if isinstance(pf, dict) else []
    if not fin_p and vin_p:
        fin_p = [vin_p]
    # Punti: 50 per ogni finalista indovinato
    fin_reali = [x for x in [casa_r, trasf_r] if x]
    punti_fin = sum(50 for sq in fin_reali if sq in (fin_p or []))
    punti_win = 70 if (vin_r and vin_p == vin_r) else 0
    # +5pt modalità finale (solo se vincitore corretto)
    mod_ok_fin = bool(vin_r and vin_p == vin_r and mod_p and mod_r and mod_p == mod_r)
    punti_mod  = 5 if mod_ok_fin else 0
    out['finale'] = {
        'F': {
            'casa_r': casa_r, 'trasf_r': trasf_r,
            'vin_prono': vin_p, 'vin_reale': vin_r,
            'mod_prono': mod_p, 'mod_reale': mod_r,
            'fin_prono': fin_p, 'fin_reali': fin_reali,
            'sq_ok': bool(fin_p and fin_reali and any(sq in fin_reali for sq in fin_p)),
            'mod_ok': mod_ok_fin,
            'punti_finalisti': punti_fin, 'punti_vincitore': punti_win, 'punti_modalita': punti_mod,
            'punti': punti_fin + punti_win + punti_mod,
        }
    }
    return out

def _calcola_capocannoniere(prono, risultati):
    rC = (risultati.get('capocannoniere_finale') or {})
    pC = (prono.get('capocannoniere') or {})
    cp1 = rC.get('primo');  cp2 = rC.get('secondo'); cp3 = rC.get('terzo')
    pp1 = pC.get('primo');  pp2 = pC.get('secondo'); pp3 = pC.get('terzo')
    terna_r = [x for x in [cp1, cp2, cp3] if x]
    terna_p = [x for x in [pp1, pp2, pp3] if x]
    punti = 0
    dettaglio = []
    if cp1 and pp1 == cp1: punti += 40; dettaglio.append('1°✓')
    if cp2 and pp2 == cp2: punti += 20; dettaglio.append('2°✓')
    if cp3 and pp3 == cp3: punti += 10; dettaglio.append('3°✓')
    # nella_terna: predicted player is in real terna but NOT at their exact position
    nella_terna = []
    for _p, _cp in [(pp1, cp1), (pp2, cp2), (pp3, cp3)]:
        if _p and _p in terna_r and _p != _cp:
            nella_terna.append(_p)
    if nella_terna:
        punti += 10; dettaglio.append('terna✓')
    return {
        'primo_prono': pp1, 'primo_reale': cp1, 'ok_1': cp1 and pp1 == cp1,
        'secondo_prono': pp2, 'secondo_reale': cp2, 'ok_2': cp2 and pp2 == cp2,
        'terzo_prono': pp3, 'terzo_reale': cp3, 'ok_3': cp3 and pp3 == cp3,
        'punti': punti, 'dettaglio': ' '.join(dettaglio),
    }

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: RISULTATI (input)
# ══════════════════════════════════════════════════════════════════════════════

# Costante: riga di inizio dei dati gironi nel sheet RISULTATI
RISULTATI_DATA_ROW = 5   # riga in cui inizia A1 (righe 1-4 = titolo + intestazioni)
RISULTATI_GC_COL  = 5   # colonna E
RISULTATI_GT_COL  = 6   # colonna F
RISULTATI_SG_COL  = 7   # colonna G (formula)

def _build_risultati(ws, db, risultati):
    ws.freeze_panes = 'A5'

    # Titolo
    _merge(ws, 1, 1, 1, 9)
    c = ws.cell(1, 1, '🏆 MONDIALITO 2026 — Risultati Ufficiali (INSERISCI QUI)')
    _style(ws, c, H_TITLE); c.alignment = _align('center')
    _set_row_height(ws, 1, 32)

    # Istruzione
    _merge(ws, 2, 1, 2, 9)
    c = ws.cell(2, 1, '🟡 Celle gialle = inserisci il risultato ufficiale · 🟢 Celle verdi = calcolate automaticamente')
    _style(ws, c, {'fill': _fill(GIALLO), 'font': _font(italic=True, color='444444', size=10), 'align': _align('center')})
    _set_row_height(ws, 2, 20)

    # Header colonne
    headers = ['ID', 'Gir.', 'Casa', 'Trasferta', 'Gol Casa ★', 'Gol Trasf. ★', 'Segno', 'Data', 'Stadio']
    widths  = [6, 5, 18, 18, 12, 12, 8, 12, 22]
    _set_row_height(ws, 3, 10)

    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(4, j, h)
        _style(ws, c, H_COL)
        _set_col_width(ws, j, w)
    _set_row_height(ws, 4, 22)

    # Dati gironi
    rg = risultati.get('gironi', {})
    match_order = []
    for lettera, girone in sorted(db['gironi'].items()):
        for partita in girone['partite']:
            match_order.append((lettera, partita))

    squadre_map = {s['id']: s['nome'] for s in db['squadre']}

    for i, (lettera, p) in enumerate(match_order):
        row = RISULTATI_DATA_ROW + i
        r   = rg.get(p['id'], {})
        gc  = r.get('gol_casa')
        gt  = r.get('gol_trasferta')

        data_row = [
            p['id'], lettera,
            squadre_map.get(p['casa'].upper(), p['casa']),
            squadre_map.get(p['trasferta'].upper(), p['trasferta']),
            gc, gt,
            None,   # segno formula
            p.get('data', ''), p.get('stadio', ''),
        ]

        alt = (i % 2 == 0)
        bg  = 'F0F5FB' if alt else BIANCO

        for j, val in enumerate(data_row, 1):
            c = ws.cell(row, j)
            if j in (1, 2):   # ID, Girone
                c.value = val
                c.fill  = _fill(bg); c.font = _font(bold=True, size=10); c.alignment = _align('center')
            elif j in (3, 4): # squadre
                c.value = val
                c.fill  = _fill(bg); c.font = _font(size=10); c.alignment = _align('left')
            elif j in (5, 6): # GOL — input giallo
                c.value = val
                c.fill  = _fill(GIALLO if val is None else 'FFF9C4')
                c.font  = _font(bold=True, size=12); c.alignment = _align('center')
                c.number_format = '0'
            elif j == 7:      # SEGNO — formula
                ec = get_column_letter(RISULTATI_GC_COL)
                fc = get_column_letter(RISULTATI_GT_COL)
                c.value = (f'=IF({ec}{row}="","",IF({ec}{row}>{fc}{row},"1",'
                           f'IF({ec}{row}<{fc}{row},"2","X")))')
                c.fill  = _fill(VERDE_CHIARO); c.font = _font(bold=True, size=11); c.alignment = _align('center')
            else:
                c.value = val
                c.fill  = _fill(bg); c.font = _font(size=9, italic=True, color='666666')
                c.alignment = _align('center')
            c.border = _border()

    # Separatore + header sezione eliminatoria
    elim_start_row = RISULTATI_DATA_ROW + 72 + 2
    _merge(ws, elim_start_row, 1, elim_start_row, 9)
    c = ws.cell(elim_start_row, 1, '🏟️  FASE ELIMINATORIA — Inserisci vincitori e modalità (90min / supplementari / rigori)')
    _style(ws, c, H_PHASE); c.alignment = _align('left')
    _set_row_height(ws, elim_start_row, 24)

    elim_hdr = elim_start_row + 1
    elim_hdr_list = ['ID', 'Fase', 'Casa (stima)', 'Trasferta (stima)', 'Vincitore ★', 'Modalità ★']
    elim_hdr_w    = [6, 10, 18, 18, 18, 14]
    for j, (h, w) in enumerate(zip(elim_hdr_list, elim_hdr_w), 1):
        c = ws.cell(elim_hdr, j, h); _style(ws, c, H_COL); _set_col_width(ws, j, w)
    _set_row_height(ws, elim_hdr, 20)

    fasi_elim = [
        ('sedicesimi', 'Sedic.', [f'S{i:02d}' for i in range(1, 17)]),
        ('ottavi',     'Ottavi', [f'O{i}'     for i in range(1, 9)]),
        ('quarti',     'Quarti', [f'Q{i}'     for i in range(1, 5)]),
        ('semifinali', 'Semi',   ['SF1', 'SF2']),
        ('finale',     'Finale', ['F']),
    ]
    rElim = risultati.get('fase_eliminatoria', {})
    erow = elim_hdr + 1
    for fase_key, fase_label, match_ids in fasi_elim:
        rFase = rElim.get(fase_key, {})
        for mid in match_ids:
            rm = rFase.get(mid, {}) or {}
            alt = (erow % 2 == 0)
            bg  = 'F0F5FB' if alt else BIANCO
            vals = [mid, fase_label,
                    rm.get('casa', ''), rm.get('trasferta', ''),
                    rm.get('vincitore', ''), rm.get('modalita', '')]
            for j, v in enumerate(vals, 1):
                c = ws.cell(erow, j, v)
                if j in (5, 6):
                    c.fill = _fill(GIALLO if not v else 'FFF9C4')
                    c.font = _font(bold=(j == 5), size=11 if j == 5 else 10)
                else:
                    c.fill = _fill(bg); c.font = _font(size=10, italic=(j == 3 or j == 4), color='666666' if j > 2 else '000000')
                c.alignment = _align('center')
                c.border = _border()
            erow += 1

    # Capocannoniere
    erow += 1
    _merge(ws, erow, 1, erow, 6)
    c = ws.cell(erow, 1, '👟 CAPOCANNONIERE — Inserisci i 3 classificati ufficiali')
    _style(ws, c, H_PHASE); c.alignment = _align('left')
    _set_row_height(ws, erow, 24)
    erow += 1
    rC = risultati.get('capocannoniere_finale', {})
    for pos_label, key in [('1° Cannoniere ★', 'primo'), ('2° Cannoniere ★', 'secondo'), ('3° Cannoniere ★', 'terzo')]:
        val = rC.get(key, '')
        c1 = ws.cell(erow, 1, pos_label); _style(ws, c1, LABEL_CELL); c1.border = _border()
        c2 = ws.cell(erow, 2, val)
        c2.fill = _fill(GIALLO if not val else 'FFF9C4'); c2.font = _font(bold=True, size=11)
        c2.alignment = _align('center'); c2.border = _border()
        for j in range(3, 7): ws.cell(erow, j).border = _border()
        erow += 1

    # Posizioni finali gironi
    erow += 1
    _merge(ws, erow, 1, erow, 7)
    c = ws.cell(erow, 1, '📊 POSIZIONI FINALI NEI GIRONI — 1° · 2° · 3° · 4°')
    _style(ws, c, H_PHASE); c.alignment = _align('left')
    _set_row_height(ws, erow, 24)
    erow += 1
    hdr_pos = ['Girone', '1° ★', '2° ★', '3° ★', '4° ★']
    for j, h in enumerate(hdr_pos, 1):
        c = ws.cell(erow, j, h); _style(ws, c, H_COL); c.border = _border()
    _set_row_height(ws, erow, 20)
    erow += 1
    rGriglia = risultati.get('posizioni_finali_gironi', {})
    for lettera in sorted(db['gironi'].keys()):
        pos = rGriglia.get(lettera, ['', '', '', ''])
        while len(pos) < 4: pos.append('')
        c = ws.cell(erow, 1, f'Girone {lettera}')
        _style(ws, c, LABEL_CELL); c.border = _border()
        for j, sq in enumerate(pos[:4], 2):
            c = ws.cell(erow, j, sq)
            c.fill = _fill(GIALLO if not sq else 'FFF9C4')
            c.font = _font(bold=True, size=10); c.alignment = _align('center'); c.border = _border()
        erow += 1

    ws.sheet_view.showGridLines = True
    ws.print_title_rows = '1:4'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: GIRONI (pronostici + formule punti)
# ══════════════════════════════════════════════════════════════════════════════

GIRONI_FIXED_COLS = 7   # A..G
GIRONI_USER_COLS  = 4   # GC, GT, Segno, Punti per utente
GIRONI_DATA_ROW   = 4   # prima riga dati (righe 1-3 = titolo, user header, col header)

def _build_gironi(ws, db, users, risultati):
    ws.freeze_panes = f'{get_column_letter(GIRONI_FIXED_COLS + 1)}4'
    n = len(users)

    # Costruisci ordine partite
    match_order = []
    for lettera, girone in sorted(db['gironi'].items()):
        for p in girone['partite']:
            match_order.append((lettera, p))
    total_rows = len(match_order)   # 72

    squadre_map  = {s['id']: s['nome'] for s in db['squadre']}
    squadre_flag = {s['id']: s.get('flag', '') for s in db['squadre']}

    # ── Titolo (riga 1) ──────────────────────────────────
    total_cols = GIRONI_FIXED_COLS + n * GIRONI_USER_COLS
    _merge(ws, 1, 1, 1, max(total_cols, GIRONI_FIXED_COLS + 1))
    c = ws.cell(1, 1, '⚽ FASE A GIRONI — Pronostici e punteggi')
    _style(ws, c, H_TITLE)
    _set_row_height(ws, 1, 28)

    # ── Header utenti (riga 2) ────────────────────────────
    # Colonne fisse
    for j in range(1, GIRONI_FIXED_COLS + 1):
        c = ws.cell(2, j); c.fill = _fill(BLU_CHIARO)
    _merge(ws, 2, 1, 2, GIRONI_FIXED_COLS)
    c = ws.cell(2, 1, 'RISULTATO UFFICIALE')
    _style(ws, c, H_COL); c.alignment = _align('center')
    _set_row_height(ws, 2, 22)

    for i, user in enumerate(users):
        col_start = GIRONI_FIXED_COLS + 1 + i * GIRONI_USER_COLS
        _merge(ws, 2, col_start, 2, col_start + GIRONI_USER_COLS - 1)
        c = ws.cell(2, col_start, user['nome'])
        _style(ws, c, H_USER)

    # ── Sub-header colonne (riga 3) ───────────────────────
    fixed_hdrs  = ['ID', 'Gir.', 'Casa', 'Trasferta', 'Gol C.', 'Gol T.', 'Segno']
    fixed_widths = [6, 5, 18, 18, 7, 7, 7]
    for j, (h, w) in enumerate(zip(fixed_hdrs, fixed_widths), 1):
        c = ws.cell(3, j, h); _style(ws, c, H_COL); _set_col_width(ws, j, w)
    user_sub = ['G.C.', 'G.T.', 'Sgn', 'Pt']
    user_sub_w = [5, 5, 5, 6]
    for i in range(n):
        for k, (h, w) in enumerate(zip(user_sub, user_sub_w)):
            col = GIRONI_FIXED_COLS + 1 + i * GIRONI_USER_COLS + k
            c = ws.cell(3, col, h); _style(ws, c, H_USER2); _set_col_width(ws, col, w)
    _set_row_height(ws, 3, 20)

    # ── Dati (righe 4..75) ────────────────────────────────
    for mi, (lettera, p) in enumerate(match_order):
        row   = GIRONI_DATA_ROW + mi
        ris_r = RISULTATI_DATA_ROW + mi    # riga corrispondente in RISULTATI
        alt   = (mi % 2 == 0)
        bg    = 'F0F5FB' if alt else BIANCO

        # Colonne fisse
        fixed_vals = [
            p['id'], lettera,
            f"{squadre_flag.get(p['casa'].upper(),'')} {squadre_map.get(p['casa'].upper(), p['casa'])}",
            f"{squadre_flag.get(p['trasferta'].upper(),'')} {squadre_map.get(p['trasferta'].upper(), p['trasferta'])}",
        ]
        for j, v in enumerate(fixed_vals, 1):
            c = ws.cell(row, j, v)
            c.fill = _fill(bg)
            c.font = _font(bold=(j <= 2), size=10)
            c.alignment = _align('center' if j <= 2 else 'left')
            c.border = _border()

        # Gol ufficiali (formula da RISULTATI)
        gc_col = get_column_letter(RISULTATI_GC_COL)
        gt_col = get_column_letter(RISULTATI_GT_COL)
        sg_col = get_column_letter(RISULTATI_SG_COL)

        for j, formula in [
            (5, f"='RISULTATI'!{gc_col}{ris_r}"),
            (6, f"='RISULTATI'!{gt_col}{ris_r}"),
            (7, f"='RISULTATI'!{sg_col}{ris_r}"),
        ]:
            c = ws.cell(row, j, formula)
            c.fill  = _fill(VERDE_CHIARO); c.font = _font(bold=(j == 7), size=10 if j < 7 else 11)
            c.alignment = _align('center'); c.border = _border()

        # Colonne per ogni utente
        segno_uff_cell = f'G{row}'   # segno ufficiale in questo sheet
        for i, user in enumerate(users):
            col_gc = GIRONI_FIXED_COLS + 1 + i * GIRONI_USER_COLS
            col_gt = col_gc + 1
            col_sg = col_gc + 2
            col_pt = col_gc + 3

            pg  = user['pronostici'].get('gironi', {})
            pm  = pg.get(p['id'], {})
            gc_p = pm.get('gol_casa');   gt_p = pm.get('gol_trasferta'); sg_p = pm.get('segno', '')

            # GC, GT, Segno — valori statici dall'export
            for j, v in [(col_gc, gc_p), (col_gt, gt_p), (col_sg, sg_p)]:
                c = ws.cell(row, j)
                c.value = v if v is not None else ''
                c.fill  = _fill(bg if v is None or v == '' else 'EBF3FB')
                c.font  = _font(size=10, bold=(j == col_sg))
                c.alignment = _align('center'); c.border = _border()

            # Punti — formula Excel
            gc_ltr = get_column_letter(col_gc)
            gt_ltr = get_column_letter(col_gt)
            sg_ltr = get_column_letter(col_sg)
            ris_gc = f"'RISULTATI'!{gc_col}{ris_r}"
            ris_gt = f"'RISULTATI'!{gt_col}{ris_r}"

            formula_punti = (
                f'=IF({segno_uff_cell}="","",IF({sg_ltr}{row}={segno_uff_cell},'
                f'10+IF(AND({gc_ltr}{row}={ris_gc},{gt_ltr}{row}={ris_gt}),5,0),0))'
            )
            c = ws.cell(row, col_pt, formula_punti)
            c.fill  = _fill(VERDE_CHIARO); c.font = _font(bold=True, size=10, color=VERDE_SCURO)
            c.alignment = _align('center'); c.border = _border()
            c.number_format = '0'

    # ── Riga TOTALE ────────────────────────────────────────
    tot_row = GIRONI_DATA_ROW + total_rows
    _merge(ws, tot_row, 1, tot_row, GIRONI_FIXED_COLS)
    c = ws.cell(tot_row, 1, 'TOTALE GIRONI (punti per utente)')
    _style(ws, c, H_COL); c.alignment = _align('right')
    _set_row_height(ws, tot_row, 24)

    for i in range(n):
        col_pt = GIRONI_FIXED_COLS + 4 + i * GIRONI_USER_COLS
        c_ltr  = get_column_letter(col_pt)
        c = ws.cell(tot_row, col_pt,
                    f'=SUM({c_ltr}{GIRONI_DATA_ROW}:{c_ltr}{GIRONI_DATA_ROW + total_rows - 1})')
        c.fill  = _fill(BLU_SCURO); c.font = _font(bold=True, color=BIANCO, size=12)
        c.alignment = _align('center'); c.border = _border()

    ws.sheet_view.showGridLines = True
    ws.print_title_rows = '1:3'
    ws.print_title_cols = 'A:G'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: GRIGLIA (posizioni finali gironi)
# ══════════════════════════════════════════════════════════════════════════════

def _build_griglia(ws, db, users, risultati):
    ws.freeze_panes = 'D4'
    n = len(users)
    squadre_map  = {s['id']: s['nome'] for s in db['squadre']}
    squadre_flag = {s['id']: s.get('flag', '') for s in db['squadre']}

    # Titolo
    tot_cols = 3 + n * 2
    _merge(ws, 1, 1, 1, max(tot_cols, 5))
    c = ws.cell(1, 1, '📊 POSIZIONI FINALI NEI GIRONI — Pronostici e punteggi (10 pt per posizione corretta)')
    _style(ws, c, H_TITLE); _set_row_height(ws, 1, 28)

    # Header utenti (riga 2)
    _merge(ws, 2, 1, 2, 3)
    c = ws.cell(2, 1, 'INFO'); _style(ws, c, H_COL)
    for i, user in enumerate(users):
        col = 4 + i * 2
        _merge(ws, 2, col, 2, col + 1)
        c = ws.cell(2, col, user['nome']); _style(ws, c, H_USER)
    _set_row_height(ws, 2, 22)

    # Sub-header (riga 3)
    for j, h in enumerate(['Gir.', 'Squadra', 'Pos. Ufficiale'], 1):
        c = ws.cell(3, j, h); _style(ws, c, H_COL)
        _set_col_width(ws, j, 6 if j == 1 else (20 if j == 2 else 14))
    for i in range(n):
        for k, h in enumerate(['Prono', 'Punti'], 1):
            col = 4 + i * 2 + k - 1
            c = ws.cell(3, col, h); _style(ws, c, H_USER2)
            _set_col_width(ws, col, 7)
    _set_row_height(ws, 3, 20)

    rGriglia = risultati.get('posizioni_finali_gironi', {})
    row = 4
    for lettera, girone in sorted(db['gironi'].items()):
        sq_list = girone['squadre']
        pos_reale = rGriglia.get(lettera, [])

        for sq in sq_list:
            alt = (row % 2 == 0)
            bg  = 'F0F5FB' if alt else BIANCO
            pr  = pos_reale.index(sq) + 1 if sq in pos_reale else None

            c = ws.cell(row, 1, lettera)
            c.fill = _fill(bg); c.font = _font(bold=True, size=10); c.alignment = _align('center'); c.border = _border()

            nome_sq = f"{squadre_flag.get(sq,'')} {squadre_map.get(sq, sq)}"
            c = ws.cell(row, 2, nome_sq)
            c.fill = _fill(bg); c.font = _font(size=10); c.alignment = _align('left'); c.border = _border()

            c = ws.cell(row, 3, pr if pr else '—')
            c.fill = _fill(VERDE_CHIARO if pr else bg)
            c.font = _font(bold=bool(pr), size=10); c.alignment = _align('center'); c.border = _border()

            for i, user in enumerate(users):
                pg   = user['pronostici'].get('posizioni_girone') or _derive_posizioni_girone(user['pronostici'], db)
                pos  = pg.get(lettera, [])
                pp   = pos.index(sq) + 1 if sq in pos else None
                ok   = (pp is not None and pr is not None and pp == pr)

                col_p = 4 + i * 2
                col_s = col_p + 1

                c = ws.cell(row, col_p, pp if pp else '—')
                c.fill = _fill(bg if not pp else 'EBF3FB')
                c.font = _font(size=10); c.alignment = _align('center'); c.border = _border()

                punti = 10 if ok else 0
                c = ws.cell(row, col_s, punti if punti else '')
                c.fill  = _fill(VERDE_CHIARO if punti else GRIGIO_CH)
                c.font  = _font(bold=bool(punti), size=10, color=VERDE_SCURO if punti else 'AAAAAA')
                c.alignment = _align('center'); c.border = _border()

            row += 1

        # Riga separatore tra gironi
        for j in range(1, 4 + n * 2 + 1):
            c = ws.cell(row, j)
            c.fill = _fill(BLU_CHIARO); c.border = _border()
        ws.row_dimensions[row].height = 4
        row += 1

    # Totali per utente
    tot_row = row + 1
    _merge(ws, tot_row, 1, tot_row, 3)
    c = ws.cell(tot_row, 1, 'TOTALE GRIGLIA')
    _style(ws, c, H_COL); c.alignment = _align('right')
    _set_row_height(ws, tot_row, 24)

    for i, user in enumerate(users):
        col_s = 5 + i * 2
        # Somma tutte le celle punti di questo utente
        total = 0
        for lettera, girone in sorted(db['gironi'].items()):
            for sq in girone['squadre']:
                rg = risultati.get('posizioni_finali_gironi', {})
                pos_reale = rg.get(lettera, [])
                pg = user['pronostici'].get('posizioni_girone') or _derive_posizioni_girone(user['pronostici'], db)
                pos = pg.get(lettera, [])
                pp = pos.index(sq) + 1 if sq in pos else None
                pr = pos_reale.index(sq) + 1 if sq in pos_reale else None
                if pp and pr and pp == pr:
                    total += 10
        c = ws.cell(tot_row, col_s, total)
        c.fill = _fill(BLU_SCURO); c.font = _font(bold=True, color=BIANCO, size=12)
        c.alignment = _align('center'); c.border = _border()

    ws.sheet_view.showGridLines = True

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: ELIMINATORIE
# ══════════════════════════════════════════════════════════════════════════════

def _build_eliminatorie(ws, db, users, risultati):
    ws.freeze_panes = 'E4'
    n = len(users)
    squadre_map = {s['id']: s.get('nome', s['id']) for s in db['squadre']}

    fasi_config = [
        ('sedicesimi', 'Sedicesimi', [f'S{i:02d}' for i in range(1, 17)], 5,  VIOLA),
        ('ottavi',     'Ottavi',     [f'O{i}'     for i in range(1, 9)],  10, BLU_MEDIO),
        ('quarti',     'Quarti',     [f'Q{i}'     for i in range(1, 5)],  20, '375623'),
        ('semifinali', 'Semifinali', ['SF1','SF2'],                        30, ROSSO),
        ('finale',     'Finale',     ['F'],                               50, '7B3F00'),
    ]

    # Titolo
    tot_cols = 4 + n * 3
    _merge(ws, 1, 1, 1, max(tot_cols, 6))
    c = ws.cell(1, 1, '🏟️ FASE ELIMINATORIA — Pronostici e punteggi')
    _style(ws, c, H_TITLE); _set_row_height(ws, 1, 28)

    row = 3
    rElim = risultati.get('fase_eliminatoria', {})

    for fase_key, fase_label, match_ids, pti_sq, color in fasi_config:
        # Intestazione fase
        _merge(ws, row, 1, row, max(tot_cols, 6))
        c = ws.cell(row, 1,
                    f'━━  {fase_label.upper()}  ·  '
                    f'{pti_sq} pt/squadra'
                    + (f'  +5 pt modalità' if fase_key != 'finale' else '  · 50pt/finalista · 70pt vincitore'))
        c.fill  = _fill(color); c.font = _font(bold=True, color=BIANCO, size=11)
        c.alignment = _align('left'); _set_row_height(ws, row, 26); row += 1

        # Sub-header
        fixed_h = ['ID', 'Casa (reale)', 'Trasferta (reale)', 'Vincitore REALE']
        for j, h in enumerate(fixed_h, 1):
            c = ws.cell(row, j, h); _style(ws, c, H_COL)
            _set_col_width(ws, j, 7 if j == 1 else (18 if j < 4 else 18))
        for i, user in enumerate(users):
            col = 5 + i * 3
            _merge(ws, row, col, row, col + 2)
            c = ws.cell(row, col, user['nome']); _style(ws, c, H_USER)
            _set_col_width(ws, col, 14); _set_col_width(ws, col+1, 10); _set_col_width(ws, col+2, 7)
        _set_row_height(ws, row, 20); row += 1

        sub_h = ['Vincitore', 'Mod.', 'Pt']
        for i in range(n):
            for k, h in enumerate(sub_h):
                c = ws.cell(row, 5 + i*3 + k, h); _style(ws, c, H_USER2)
        _set_row_height(ws, row, 18); row += 1

        rFase = rElim.get(fase_key, {})

        for mid in match_ids:
            rm   = rFase.get(mid, {}) or {}
            alt  = (row % 2 == 0)
            bg   = 'F5F0FB' if alt else BIANCO

            # Colonne fisse
            vals_f = [mid,
                      squadre_map.get((rm.get('casa') or '').upper(), rm.get('casa', '—')),
                      squadre_map.get((rm.get('trasferta') or '').upper(), rm.get('trasferta', '—')),
                      squadre_map.get((rm.get('vincitore') or '').upper(), rm.get('vincitore', '—'))]
            for j, v in enumerate(vals_f, 1):
                c = ws.cell(row, j, v)
                c.fill = _fill(bg if j < 4 else VERDE_CHIARO)
                c.font = _font(bold=(j == 1 or j == 4), size=10)
                c.alignment = _align('center'); c.border = _border()

            # Colonne utenti
            for i, user in enumerate(users):
                col   = 5 + i * 3
                pElim = user['pronostici'].get('fase_eliminatoria', {})
                pFase = pElim.get(fase_key, {})
                pm    = pFase.get(mid, {}) or {}

                if fase_key == 'finale':
                    # per la finale: vincitore dal campo vincitore (struttura Firestore: F.vincitore)
                    pf = pm if isinstance(pm, dict) else {}
                    vin_p = pf.get('vincitore')
                    fin_p = pf.get('squadre') or ([vin_p] if vin_p else [])
                    mod_p = pf.get('modalita', '')
                    vin_r = rm.get('vincitore')
                    mod_r = rm.get('modalita', '')
                    fin_r = [rm.get('casa'), rm.get('trasferta')]

                    punti_fin = sum(50 for sq in [x for x in fin_r if x] if sq in (fin_p or []))
                    punti_win = 70 if (vin_r and vin_p == vin_r) else 0
                    punti_mod = 5 if (vin_r and vin_p == vin_r and mod_p and mod_r and mod_p == mod_r) else 0
                    punti     = punti_fin + punti_win + punti_mod

                    disp_v = squadre_map.get((vin_p or '').upper(), vin_p or '—')
                    disp_m = mod_p or '—'
                else:
                    vin_p = pm.get('vincitore', '')
                    mod_p = pm.get('modalita', '')
                    vin_r = rm.get('vincitore')
                    sq_reali = set()
                    for m2 in rFase.values():
                        if isinstance(m2, dict):
                            for f in ('casa', 'trasferta', 'vincitore'):
                                if m2.get(f): sq_reali.add(m2[f])
                    sq_ok  = bool(vin_p and sq_reali and vin_p in sq_reali)
                    mod_ok = bool(sq_ok and vin_p == vin_r and mod_p and rm.get('modalita') and mod_p == rm.get('modalita'))
                    punti  = (pti_sq if sq_ok else 0) + (5 if mod_ok else 0)

                    disp_v = squadre_map.get((vin_p or '').upper(), vin_p or '—')
                    disp_m = mod_p or '—'

                c = ws.cell(row, col, disp_v)
                ok = (vin_p and rm.get('vincitore') and vin_p == rm.get('vincitore'))
                c.fill  = _fill(VERDE_CHIARO if ok else bg)
                c.font  = _font(bold=ok, size=10); c.alignment = _align('center'); c.border = _border()

                c = ws.cell(row, col+1, disp_m)
                c.fill  = _fill(bg); c.font = _font(size=9, italic=True, color='555555')
                c.alignment = _align('center'); c.border = _border()

                c = ws.cell(row, col+2, punti if punti else '')
                c.fill  = _fill(VERDE_CHIARO if punti else GRIGIO_CH)
                c.font  = _font(bold=bool(punti), size=10, color=VERDE_SCURO if punti else 'AAAAAA')
                c.alignment = _align('center'); c.border = _border()

            row += 1

        # Riga totale per fase
        _merge(ws, row, 1, row, 4)
        c = ws.cell(row, 1, f'Totale {fase_label}')
        _style(ws, c, H_COL); c.alignment = _align('right')
        _set_row_height(ws, row, 20)

        for i in range(n):
            col_pt = 5 + i * 3 + 2   # colonna punti per utente i
            fase_pt = 0
            pElim = users[i]['pronostici'].get('fase_eliminatoria', {})
            pFase = pElim.get(fase_key, {})
            sq_reali = set()
            for m2 in rFase.values():
                if isinstance(m2, dict):
                    for f in ('casa', 'trasferta', 'vincitore'):
                        if m2.get(f): sq_reali.add(m2[f])
            for mid in match_ids:
                rm = rFase.get(mid, {}) or {}
                pm = pFase.get(mid, {}) or {}
                if fase_key == 'finale':
                    vin_p = pm.get('vincitore') if isinstance(pm, dict) else None
                    fin_p = (pm.get('squadre') if isinstance(pm, dict) else None) or ([vin_p] if vin_p else [])
                    vin_r = rm.get('vincitore')
                    mod_p = pm.get('modalita') if isinstance(pm, dict) else None
                    mod_r = rm.get('modalita')
                    fin_r = [rm.get('casa'), rm.get('trasferta')]
                    fase_pt += sum(50 for sq in [x for x in fin_r if x] if sq in (fin_p or []))
                    fase_pt += 70 if (vin_r and vin_p == vin_r) else 0
                    fase_pt += 5 if (vin_r and vin_p == vin_r and mod_p and mod_r and mod_p == mod_r) else 0
                else:
                    vin_p = pm.get('vincitore', '')
                    if vin_p and sq_reali and vin_p in sq_reali:
                        fase_pt += pti_sq
                        if vin_p == rm.get('vincitore') and pm.get('modalita') == rm.get('modalita'):
                            fase_pt += 5
            c = ws.cell(row, col_pt, fase_pt if fase_pt else '')
            c.fill = _fill(BLU_SCURO if fase_pt else GRIGIO_CH)
            c.font = _font(bold=True, color=BIANCO if fase_pt else 'AAAAAA', size=11)
            c.alignment = _align('center'); c.border = _border()

        row += 2  # spazio tra fasi

    ws.sheet_view.showGridLines = True

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: CAPOCANNONIERE
# ══════════════════════════════════════════════════════════════════════════════

def _build_capocannoniere(ws, db, users, risultati):
    ws.freeze_panes = 'C4'
    n = len(users)
    rC = risultati.get('capocannoniere_finale', {})
    cp = [rC.get('primo','—'), rC.get('secondo','—'), rC.get('terzo','—')]

    _merge(ws, 1, 1, 1, max(3 + n * 2, 5))
    c = ws.cell(1, 1, '👟 CAPOCANNONIERE — Pronostici e punteggi')
    _style(ws, c, H_TITLE); _set_row_height(ws, 1, 28)

    # Header
    _merge(ws, 2, 1, 2, 2)
    c = ws.cell(2, 1, 'POSIZIONE'); _style(ws, c, H_COL)
    _set_col_width(ws, 1, 14); _set_col_width(ws, 2, 22)
    for i, user in enumerate(users):
        col = 3 + i * 2
        _merge(ws, 2, col, 2, col + 1)
        c = ws.cell(2, col, user['nome']); _style(ws, c, H_USER)
        _set_col_width(ws, col, 22); _set_col_width(ws, col+1, 7)
    _set_row_height(ws, 2, 22)

    sub = ['Reale']
    for j, h in enumerate(sub, 2):
        c = ws.cell(3, j, h); _style(ws, c, H_COL)
    for i in range(n):
        for k, h in enumerate(['Prono', 'Pt']):
            c = ws.cell(3, 3 + i*2 + k, h); _style(ws, c, H_USER2)
    _set_row_height(ws, 3, 18)

    pos_info = [
        ('🥇 1° Cannoniere', 'primo',  40),
        ('🥈 2° Cannoniere', 'secondo', 20),
        ('🥉 3° Cannoniere', 'terzo',   10),
    ]

    for row, (label, key, pts_exact) in enumerate(pos_info, 4):
        alt  = (row % 2 == 0)
        bg   = 'F0F5FB' if alt else BIANCO
        real = rC.get(key, '—')

        c = ws.cell(row, 1, label)
        c.fill = _fill(bg); c.font = _font(bold=True, size=11); c.alignment = _align('left'); c.border = _border()

        c = ws.cell(row, 2, real)
        c.fill = _fill(VERDE_CHIARO if real != '—' else bg)
        c.font = _font(bold=True, size=11); c.alignment = _align('center'); c.border = _border()

        for i, user in enumerate(users):
            col_p = 3 + i * 2
            col_s = col_p + 1
            pC    = user['pronostici'].get('capocannoniere', {})
            pp    = pC.get(key, '—')
            terna_r = [rC.get('primo'), rC.get('secondo'), rC.get('terzo')]
            ok_exact = (real != '—' and pp == real)
            ok_terna = (not ok_exact and pp in terna_r)
            punti = pts_exact if ok_exact else (10 if ok_terna else 0)

            c = ws.cell(row, col_p, pp or '—')
            c.fill = _fill(VERDE_CHIARO if ok_exact else ('FFF2CC' if ok_terna else bg))
            c.font = _font(bold=ok_exact, size=10); c.alignment = _align('center'); c.border = _border()

            c = ws.cell(row, col_s, punti if punti else '')
            c.fill = _fill(VERDE_CHIARO if punti else GRIGIO_CH)
            c.font = _font(bold=bool(punti), size=10, color=VERDE_SCURO if punti else 'AAAAAA')
            c.alignment = _align('center'); c.border = _border()

    # Totale per utente
    tot_row = 4 + len(pos_info) + 1
    _merge(ws, tot_row, 1, tot_row, 2)
    c = ws.cell(tot_row, 1, 'TOTALE CAPOCANNONIERE')
    _style(ws, c, H_COL); c.alignment = _align('right')

    for i, user in enumerate(users):
        col_s = 4 + i * 2
        pC    = user['pronostici'].get('capocannoniere', {})
        terna_r = [rC.get('primo'), rC.get('secondo'), rC.get('terzo')]
        tot   = 0
        for key, pts_exact in [('primo',40),('secondo',20),('terzo',10)]:
            pp = pC.get(key)
            real = rC.get(key)
            if pp and real and pp == real: tot += pts_exact
            elif pp and pp in terna_r:    tot += 10
        c = ws.cell(tot_row, col_s, tot)
        c.fill = _fill(BLU_SCURO); c.font = _font(bold=True, color=BIANCO, size=12)
        c.alignment = _align('center'); c.border = _border()

    ws.sheet_view.showGridLines = True

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: CLASSIFICA
# ══════════════════════════════════════════════════════════════════════════════

def _build_classifica(ws, db, users, risultati):
    n = len(users)

    _merge(ws, 1, 1, 1, 15)
    c = ws.cell(1, 1, '🏅 CLASSIFICA — Riepilogo punteggi Mondialito 2026')
    _style(ws, c, H_TITLE); _set_row_height(ws, 1, 32)

    _merge(ws, 2, 1, 2, 15)
    c = ws.cell(2, 1,
                '✅ Gironi = formule live (aggiornano al cambio risultati) · '
                '📌 Altri valori = calcolati al momento dell\'esportazione')
    c.fill = _fill(GIALLO); c.font = _font(italic=True, size=10, color='444444')
    c.alignment = _align('center'); _set_row_height(ws, 2, 20)

    headers = ['Pos.', 'Nome', 'Gironi', 'Griglia', 'Sedic.', 'Ottavi',
               'Quarti', 'Semi', 'Finale', 'Campione', 'Modalità', 'Cannon.', 'TOTALE']
    widths   = [6, 22, 9, 8, 8, 8, 8, 8, 8, 10, 9, 10, 10]
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(3, j, h); _style(ws, c, H_COL); _set_col_width(ws, j, w)
    _set_row_height(ws, 3, 24)

    # Calcola tutti i punteggi in Python
    scores = []
    rElim  = risultati.get('fase_eliminatoria', {})

    for user in users:
        pr = user['pronostici']
        res = risultati

        # Gironi: sarà formula Excel, ma calcoliamo per ordinare
        gironi_pt = 0
        rg = res.get('gironi', {})
        for girone in db['gironi'].values():
            for p in girone['partite']:
                r = rg.get(p['id'], {}); pm = (pr.get('gironi') or {}).get(p['id'], {})
                gc_r = r.get('gol_casa'); gt_r = r.get('gol_trasferta')
                if gc_r is None: continue
                sgn_r = _segno(gc_r, gt_r); sgn_p = pm.get('segno')
                if sgn_p == sgn_r:
                    gironi_pt += 10
                    if pm.get('gol_casa') == gc_r and pm.get('gol_trasferta') == gt_r:
                        gironi_pt += 5

        # Griglia
        griglia_d = _calcola_griglia(pr, res, db)
        griglia_pt = sum(v.get('punti', 0) for g in griglia_d.values() for v in g.values())

        # Eliminatorie
        elim_d = _calcola_eliminatorie(pr, res, db)
        sed_pt = sum(v.get('punti', 0) for v in elim_d.get('sedicesimi', {}).values())
        ott_pt = sum(v.get('punti', 0) for v in elim_d.get('ottavi', {}).values())
        qua_pt = sum(v.get('punti', 0) for v in elim_d.get('quarti', {}).values())
        sem_pt = sum(v.get('punti', 0) for v in elim_d.get('semifinali', {}).values())
        # Finale: punti finalisti
        fin_data  = elim_d.get('finale', {}).get('F', {})
        fin_pt    = fin_data.get('punti_finalisti', 0)
        camp_pt   = fin_data.get('punti_vincitore', 0)

        # Modalità
        mod_pt = 0
        for fase_key in ('sedicesimi', 'ottavi', 'quarti', 'semifinali', 'finale'):
            for v in elim_d.get(fase_key, {}).values():
                if isinstance(v, dict) and v.get('mod_ok'):
                    mod_pt += 5

        # Capocannoniere
        cann_d   = _calcola_capocannoniere(pr, res)
        cann_pt  = cann_d.get('punti', 0)

        totale = gironi_pt + griglia_pt + sed_pt + ott_pt + qua_pt + sem_pt + fin_pt + camp_pt + mod_pt + cann_pt

        scores.append({
            'user': user,
            'gironi': gironi_pt,
            'griglia': griglia_pt,
            'sed': sed_pt, 'ott': ott_pt, 'qua': qua_pt, 'sem': sem_pt,
            'fin': fin_pt, 'camp': camp_pt, 'mod': mod_pt, 'cann': cann_pt,
            'totale': totale,
        })

    # Ordina per totale desc
    scores.sort(key=lambda x: -x['totale'])

    # La colonna "Gironi" nel CLASSIFICA punta alla formula live del foglio GIRONI
    # La colonna punti nel foglio GIRONI per utente i è alla colonna 11 + i*4
    # Ma dopo l'ordinamento l'utente i nella classifica non è più in posizione i
    # → Usiamo SUMIF sul foglio GIRONI cercando il nome utente (più robusto)
    # Oppure usiamo un VLOOKUP, o semplicemente mettiamo il valore Python e
    # aggiungiamo una nota che diventerà live una volta salvato.

    medals = {1: '🥇', 2: '🥈', 3: '🥉'}

    for pos, sc in enumerate(scores, 1):
        row  = 3 + pos
        user = sc['user']

        # Trova l'indice di questo utente nella lista originale (per la formula GIRONI)
        orig_idx = next((i for i, u in enumerate(users) if u['uid'] == user['uid']), 0)
        gironi_col_letter = get_column_letter(GIRONI_FIXED_COLS + 4 + orig_idx * GIRONI_USER_COLS)
        tot_gironi_row    = GIRONI_DATA_ROW + 72

        alt  = (pos % 2 == 0)
        bg   = 'FFFDE7' if pos == 1 else ('F5F5F5' if alt else BIANCO)
        is_top = pos <= 3

        c = ws.cell(row, 1, medals.get(pos, f'{pos}°'))
        c.fill = _fill(bg); c.font = _font(bold=is_top, size=12); c.alignment = _align('center'); c.border = _border()

        c = ws.cell(row, 2, user['nome'])
        c.fill = _fill(bg); c.font = _font(bold=is_top, size=11); c.alignment = _align('left'); c.border = _border()

        # Gironi: formula live
        c = ws.cell(row, 3, f"='GIRONI'!{gironi_col_letter}{tot_gironi_row}")
        c.fill = _fill(VERDE_CHIARO); c.font = _font(bold=True, size=10, color=VERDE_SCURO)
        c.alignment = _align('center'); c.border = _border(); c.number_format = '0'

        # Valori Python per le altre categorie
        for col, val in enumerate([
            sc['griglia'], sc['sed'], sc['ott'], sc['qua'],
            sc['sem'], sc['fin'], sc['camp'], sc['mod'], sc['cann'],
        ], 4):
            c = ws.cell(row, col, val if val else '')
            c.fill = _fill(bg if not val else VERDE_CHIARO)
            c.font = _font(size=10, color=VERDE_SCURO if val else 'AAAAAA')
            c.alignment = _align('center'); c.border = _border()

        # TOTALE = formula che somma C + D:L
        c = ws.cell(row, 13, f'=C{row}+D{row}+E{row}+F{row}+G{row}+H{row}+I{row}+J{row}+K{row}+L{row}')
        tot_fill = ORO if pos == 1 else (VERDE_CHIARO if is_top else bg)
        c.fill = _fill(tot_fill); c.font = _font(bold=True, size=13 if pos == 1 else 11)
        c.alignment = _align('center'); c.border = _border_full(); c.number_format = '0'

        _set_row_height(ws, row, 26 if pos == 1 else 22)

    # Legenda
    leg_row = 3 + len(scores) + 2
    _merge(ws, leg_row, 1, leg_row, 13)
    c = ws.cell(leg_row, 1,
                '📌 Legenda punti:  Gironi: 10pt segno + 5pt bonus esatto  ·  '
                'Griglia: 10pt/pos corretta  ·  Sedic/Ott/Qua/Semi: 5/10/20/30pt/squadra  ·  '
                'Finale: 50pt/finalista  ·  Campione: 70pt  ·  Modalità: 5pt/match  ·  '
                'Cannone: 40/20/10pt pos esatta · 10pt nella terna')
    c.fill = _fill(BLU_CHIARO); c.font = _font(italic=True, size=9, color='333333')
    c.alignment = _align('left', wrap=True); _set_row_height(ws, leg_row, 36)

    ws.sheet_view.showGridLines = True
    ws.freeze_panes = 'A4'

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    script_dir = Path(__file__).parent

    # Carica DB Mondialito
    db_path = script_dir / 'mondialito_db.json'
    if not db_path.exists():
        print(f'❌ mondialito_db.json non trovato in {script_dir}')
        sys.exit(1)
    with open(db_path, encoding='utf-8') as f:
        DB = json.load(f)

    # Cerca export JSON
    export_path = None
    if len(sys.argv) > 1:
        export_path = Path(sys.argv[1])
        if not export_path.is_absolute():
            export_path = script_dir / export_path
    if not export_path or not export_path.exists():
        candidates = sorted(script_dir.glob('pronostici_mondialito*.json'), reverse=True)
        export_path = candidates[0] if candidates else None

    export = {}
    if export_path and export_path.exists():
        with open(export_path, encoding='utf-8') as f:
            export = json.load(f)
        print(f'✓ Caricato export: {export_path.name}')
    else:
        print('⚠️  Nessun file export trovato — creo Excel con struttura vuota.')
        print('   Esegui prima esporta-pronostici.js nel browser per estrarre i dati reali.')

    partecipanti   = export.get('partecipanti', {})
    pronostici_all = export.get('pronostici', {})
    risultati      = export.get('risultati', {})

    # Filtra e ordina utenti (escludi admin se vuoi, qui li includiamo)
    users = []
    for uid, p in partecipanti.items():
        nickname     = (p.get('nickname') or '').strip()
        nome_cognome = ' '.join(filter(None, [p.get('nome', ''), p.get('cognome', '')])).strip()
        nome         = nickname if nickname else (nome_cognome or uid)
        users.append({
            'uid':        uid,
            'nome':       nome,
            'pronostici': pronostici_all.get(uid, {}),
        })
    users.sort(key=lambda u: u['nome'].lower())

    if not users:
        print('⚠️  Nessun partecipante trovato — creo Excel vuoto di esempio.')
        users = [{'uid': 'esempio', 'nome': '[Esempio]', 'pronostici': {}}]

    print(f'👥 Partecipanti: {len(users)} → {", ".join(u["nome"] for u in users)}')
    n_ris = len((risultati.get('gironi') or {}))
    print(f'📊 Risultati gironi presenti: {n_ris}/72')

    # Crea workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_ris   = wb.create_sheet('RISULTATI')
    ws_gir   = wb.create_sheet('GIRONI')
    ws_gri   = wb.create_sheet('GRIGLIA')
    ws_elim  = wb.create_sheet('ELIMINATORIE')
    ws_cann  = wb.create_sheet('CAPOCANNONIERE')
    ws_class = wb.create_sheet('CLASSIFICA')

    # Tab colors
    ws_ris.sheet_properties.tabColor   = 'FFC000'
    ws_gir.sheet_properties.tabColor   = '2E6DA4'
    ws_gri.sheet_properties.tabColor   = '375623'
    ws_elim.sheet_properties.tabColor  = '7030A0'
    ws_cann.sheet_properties.tabColor  = 'C55A11'
    ws_class.sheet_properties.tabColor = '1A3A5C'

    print('🔨 Costruzione fogli...')
    _build_risultati(ws_ris,   DB, risultati)
    print('  ✓ RISULTATI')
    _build_gironi(ws_gir,      DB, users, risultati)
    print('  ✓ GIRONI')
    _build_griglia(ws_gri,     DB, users, risultati)
    print('  ✓ GRIGLIA')
    _build_eliminatorie(ws_elim, DB, users, risultati)
    print('  ✓ ELIMINATORIE')
    _build_capocannoniere(ws_cann, DB, users, risultati)
    print('  ✓ CAPOCANNONIERE')
    _build_classifica(ws_class, DB, users, risultati)
    print('  ✓ CLASSIFICA')

    # Salva
    out_path = script_dir / 'Mondialito2026_Pronostici.xlsx'
    wb.save(out_path)
    print(f'\n✅ Excel salvato: {out_path.name}')
    print(f'   Apri il file e vai su RISULTATI → inserisci i gol nelle celle gialle.')
    print(f'   Il foglio GIRONI si aggiorna automaticamente.')


if __name__ == '__main__':
    main()
